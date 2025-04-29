# views.py
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from rest_framework import viewsets, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend

from .models import Expense
from .serializers import ExpenseSerializer

class ExpenseViewSet(viewsets.ModelViewSet):
    """
    API ViewSet for managing expenses.
    Supports filtering by category, date, and expense source along with search and ordering.
    """
    queryset = Expense.objects.all().order_by("-created_at")
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["category", "expense_source", "date"]
    search_fields = ["title", "description"]
    ordering_fields = ["id", "amount", "created_at"]

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def expense_meta(request):
    """
    Returns selectable choices for category and expense_source.
    """
    # Create lists of dicts from the choices defined in your Expense model
    categories = [{"value": k, "label": v} for k, v in Expense.CATEGORY_CHOICES]
    sources = [{"value": k, "label": v} for k, v in Expense.EXPENSE_SOURCE_CHOICES]
    return Response({"categories": categories, "sources": sources})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_expenses(request):
    """
    Exports expenses to an Excel file based on applied filters.
    Accepts query parameters for:
    - category
    - source
    - start_date
    - end_date
    - search (for title/description)
    - ordering
    """
    # Extract parameters from the request
    category = request.GET.get('category')
    source = request.GET.get('source')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search = request.GET.get('search')
    ordering = request.GET.get('ordering', '-created_at')

    # Build the queryset with dynamic filters
    queryset = Expense.objects.all()

    if category:
        queryset = queryset.filter(category=category)
    if source:
        queryset = queryset.filter(expense_source=source)
    if start_date:
        queryset = queryset.filter(date__gte=start_date)
    if end_date:
        queryset = queryset.filter(date__lte=end_date)
    if search:
        queryset = queryset.filter(title__icontains=search) | queryset.filter(description__icontains=search)

    queryset = queryset.order_by(ordering)

    # Create Excel workbook and sheet using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Define header columns
    headers = ["ID", "Title", "Amount", "Category", "Source", "Date"]
    ws.append(headers)

    # Populate the sheet with expense data
    for expense in queryset:
        row = [
            expense.id,
            expense.title,
            expense.amount,
            expense.category,
            expense.expense_source,
            expense.date.strftime("%Y-%m-%d") if expense.date else ""
        ]
        ws.append(row)

    # Optionally, adjust column widths for a better view
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save workbook into an in-memory bytes buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'
    return response




